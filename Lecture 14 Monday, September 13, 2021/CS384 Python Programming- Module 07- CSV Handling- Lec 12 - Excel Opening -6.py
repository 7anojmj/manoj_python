# Add Sheets To The Excel File
from datetime import datetime
from openpyxl import load_workbook
wb = load_workbook(r"demo.xlsx")
sheet = wb.active
now = datetime.now()

date_time = now.strftime("%d.%m.%Y")
sheet.title = date_time

sheetname = "CS384 "+now.strftime("%d.%m.%Y")
wb.create_sheet(index=1, title=sheetname)
wb.create_sheet(index=2, title="SuperMan")
wb.save("demo.xlsx")
