from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
# Getting Column Letter

wb = load_workbook("sample_file.xlsx")
sheet = wb["Sheet1"]
row_count = sheet.max_row
column_count = sheet.max_column
for i in range(1, row_count + 1):
    for j in range(1, column_count + 1):
        data = sheet.cell(row=i, column=j).value
        if data == "CS384":
            column_letter = get_column_letter(j)
            print("CS384 employee name is present in column ",
                  column_letter, "and in row index ", i)
