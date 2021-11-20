# Delete A Sheet From Excel Workbook

import openpyxl
wb = openpyxl.load_workbook("DeleteSheet.xlsx")
print("The names of work sheets before deleting")
print(wb.sheetnames)
sheetDelete = wb["Sheet3"]
wb.remove(sheetDelete)  # Sheet2 will be deleted
# del wb[sheetDelete]  # Alternatively you can use del cmd
print("The names of work sheets after deleting")
print(wb.sheetnames)
wb.save("DeleteSheet.xlsx")
