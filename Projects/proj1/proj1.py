from email.mime.base import MIMEBase
from email.mime.text import MIMEText
import os
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import Workbook
import yagmail
import smtplib
from email.mime.multipart import MIMEMultipart 
from email import encoders
def generate_marksheet(path1, path2, pos, neg):
	pos=int(pos)
	neg=int(neg)
	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")
	if not os.path.isdir(os.path.join("sample_output", "marksheet")):
		os.mkdir(os.path.join("sample_output", "marksheet"))
	with open(path2, "r") as csvfile:
		global stud_info
		options_list = {}
		reader=csv.reader(csvfile)
		header=next(reader)
		head=['Roll Number','Options','Name']
		for x in reader:
			options_list[x[6]] = {}
			options_list[x[6]][head[0]] = x[6]
			options_list[x[6]][head[1]] = x[7:]
			options_list[x[6]][head[2]]=x[3]      
	if "ANSWER" not in options_list.keys():
		return r"no roll number with ANSWER is present, Cannot Process!"
    
	data={}
	corr_ans=options_list["ANSWER"]["Options"]
	for roll in options_list.keys():
		if roll!="ANSWER":
		    correct=0
		    not_attempt=0
		    wrong=0
		    data[roll] = {}
		    data[roll]["Name"] = options_list[roll]["Name"]
		    data[roll]["Roll Number"] = options_list[roll]["Roll Number"]
		    for i,val in enumerate(options_list[roll]["Options"]):
			    if val==corr_ans[i]:
				    correct+=1
			    elif val=="":
				    not_attempt += 1
			    else:
				    wrong +1
		    data[roll]["correct_no"]=correct
		    data[roll]["not_attempt_no"]=not_attempt
		    data[roll]["wrong_no"]=wrong
		    data[roll]["total_no"]=correct+not_attempt+wrong
		    data[roll]["correct_tot"]=correct*pos
		    data[roll ]["wrong_tot"]=wrong*neg
		    data[roll]["final"]=str((correct*pos)+(wrong*neg))+"/"+str((correct+not_attempt+wrong)*pos)
	for roll in data.keys():
		wb=Workbook()
		sheet=wb.active
		row_cnt=40
		col_cnt=10
		for i in range(1,row_cnt+1):
			for j in range(1,col_cnt+1):
				sheet.column_dimensions[get_column_letter(j)].width = 17
				sheet.cell(i,j).font=Font(name= "Century", size=12)
		sheet.title="quiz"
		img = openpyxl.drawing.image.Image('IITP_Logo.png')
		img.anchor = 'A1'
		sheet.add_image(img)
		sheet.merge_cells('A5:E5')
		sheet["A5"].value = 'Mark Sheet'
		sheet["A5"].font = Font(name="Century",size=18,bold=True,underline ="single")
		sheet["A5"].border = Border()
		sheet["A5"].alignment = Alignment(horizontal="center",vertical="center")
		sheet["A6"].value = "Name:"
		sheet["A6"].alignment = Alignment(horizontal="right")
		sheet["B6"].value = data[roll]["Name"]
		sheet["B6"].font = Font(name="Century",size=12,bold=True)
		sheet["B6"].alignment = Alignment(horizontal= "left")
		sheet["D6"].value = "Exam:"
		sheet["D6"].alignment = Alignment(horizontal="right")
		sheet["E6"].value = "quiz"
		sheet["E6"].font = Font(name="Century",size=12,bold=True)
		sheet["E6"].alignment = Alignment(horizontal= "left")
		sheet["A7"].value = "Roll Number:"
		sheet["A7"].alignment = Alignment(horizontal="right")
		sheet["B7"].value = data[roll]["Roll Number"]
		sheet["B7"].font = Font(name="Century",size=12,bold=True)
		sheet["B7"].alignment = Alignment(horizontal= "left")
		sheet.append(["","","","",""])
		marks_info = []
		marks_info.append(tuple(["", "Right", "Wrong", "Not Attempt", "Max"]))
		marks_info.append(tuple(["No.", data[roll]["correct_no"], data[roll]["wrong_no"],data[roll]["not_attempt_no"], data[roll]["total_no"]]))
		marks_info.append(tuple(["Marking", pos, neg, 0, ""]))
		marks_info.append(tuple(["Total", data[roll]["correct_tot"], data[roll]["wrong_tot"],"", data[roll]["final"]]))
		marks_info = tuple(marks_info)
		thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
		i1=0
		for i in range(9, 13):
			j1=0
			for j in range(1, 6):
				sheet.cell(i,j).value = marks_info[i1][j1]
				sheet.cell(i,j).border = thin_border
				sheet.cell(i,j).alignment = Alignment(horizontal= "center")
				if i1==0 or (j1==0 and i1>0):
					sheet.cell(i,j).font=Font(name="Century",size=12,bold=True)
				if i1>0:
					if j1==1:
						sheet.cell(i,j).font = Font(name="Century",size=12,color = "00008000")
					elif j1==2:
						sheet.cell(i,j).font = Font(name="Century",size=12,color="00FF0000")
				if i1==3 and j1==4:
					sheet.cell(i,j).font = Font(name="Century",size=12,color="000000FF")
				j1+=1
			i1+=1
		sheet.append(["","","","",""])
		sheet.append(["","","","",""])
		s_ans = options_list[roll]["Options"]
		c_ans = options_list["ANSWER"]["Options"]
		c1 = c_ans.copy()
		t =1
		while 1:
			if t==7:
				break
			stud_ans = []
			stud_ans.append(tuple(["Student Ans", "Correct Ans"]))
			if len(s_ans) < 25:
				f_s_ans = s_ans[:len(s_ans)]
			elif len(s_ans) >= 25:
				f_s_ans = s_ans[:25]

			if len(c1) < 25:
				f_c_ans = c1[:len(c1)]
			elif len(s_ans) >= 25:
				f_c_ans = c1[:25]
			for k in range(len(f_s_ans)):
				s_ans.pop(0)
			for k1 in range(len(f_c_ans)):
				c1.pop(0)
			for x in range(len(f_c_ans)):
				stud_ans.append(tuple([f_s_ans[x], f_c_ans[x]]))
			stud_ans = tuple(stud_ans)
			e = 15+ len(stud_ans)
			thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
			i1=0 
			for i in range(15, e):
				j1=0
				for j in range(t, t+2):
					sheet.cell(i,j).value =stud_ans[i1][j1]
					sheet.cell(i,j).border = thin_border
					sheet.cell(i,j).alignment = Alignment(horizontal= "center")
					if i1==0:
						sheet.cell(i,j).font = Font(bold = True, name= "Century", size=12)
					if i1>0:
						if (stud_ans[i1][0] == stud_ans[i1][1]):
							sheet.cell(i,j).font = Font(color = "00008000", name= "Century", size=12)
						else:
							if stud_ans[i1][0] != '':
								sheet.cell(i,j).font = Font(color = "00FF0000", name= "Century", size=12)
					if i1>0 and j1==1:
						sheet.cell(i,j).font = Font(color = "000000FF", name= "Century", size=12)
					j1+=1
				i1+=1
			t+=3

		name = os.path.join(os.getcwd(), "sample_output", "marksheet",(roll+".xlsx"))
		wb.save(name)
def concise_marksheet(path1, path2, pos, neg):
	pos=int(pos)
	neg=int(neg)
	with open(path2, "r") as csvfile:

		options_list = {}
		reader = csv.reader(csvfile)
		header = next(reader)
		head = []
		head.extend(header[0:7])
		head.append("Options")
		for row in reader:
			options_list[row[6]] = {}
			i=0
			for h in head:
				if i==7:
					break
				options_list[row[6]][h] = row[i]
				i+=1

			options_list[row[6]][head[7]] = row[7:]

	if "ANSWER" not in options_list.keys():
		return r"no roll number with ANSWER is present, Cannot Process!"
	corr_ans = options_list["ANSWER"]["Options"]
	data = {}
	for roll in options_list.keys():
		correct = 0
		not_attempt = 0
		wrong =0
		data[roll] = {}
		data[roll]["Name"] = options_list[roll]["Name"]
		data[roll]["Roll Number"] = options_list[roll]["Roll Number"]
		for i,val in enumerate(options_list[roll]["Options"]):
			if val == corr_ans[i]:
				correct += 1
			elif val == "":
				not_attempt += 1
			else:
				wrong += 1
		data[roll]["status_Ans"] = str([correct, wrong, not_attempt])
		data[roll]["score_after_neg"] = str((correct*pos)+(wrong*neg)) + "/" + str((correct+not_attempt+wrong)*pos)
		data[roll]["Score"]=str(correct*pos)+"/"+str((correct+not_attempt+wrong)*pos)

	concise_info = []
	header =["Timestamp", "Email address", "Google Score", "Name", "IITP webmail", "Phone (10 digit only)", "Score_After_Negative", "Roll Number"]
	end = 7+ (len(corr_ans)-1)

	for i in range(7, end+1):
		header.append("Unnamed: "+str(i))
	header.append("statusAns")

	for roll in options_list.keys():
		conc_data = []
		conc_data.append(options_list[roll]["Timestamp"])
		conc_data.append(options_list[roll]["Email address"])
		conc_data.append(data[roll]["Score"])
		conc_data.append(options_list[roll]["Name"])
		conc_data.append(options_list[roll]["IITP webmail"])
		conc_data.append(options_list[roll]["Phone (10 digit only)"])
		conc_data.append(data[roll]["score_after_neg"])
		conc_data.append(options_list[roll]["Roll Number"])
		conc_data.extend(options_list[roll]["Options"])
		conc_data.append(data[roll]["status_Ans"])
		concise_info.append(conc_data)
	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")
	if not os.path.isdir("sample_output//marksheet"):
		os.mkdir("sample_output//marksheet")
	with open(os.path.join("sample_output", "marksheet","concise_marksheet.csv"), "w", newline='') as f:
		csvwriter = csv.writer(f)
		csvwriter.writerow(header)
		csvwriter.writerows(concise_info)

def Send_email(path1, path2, pos, neg):
	directory=os.getcwd()
	path=directory+"/sample_output/marksheet"
	yag=yagmail.SMTP('enter your gmail here ','enter your password here')	
	with open(path2,'r') as csvfile:
		reader=csv.reader(csvfile)
		header=next(reader)
		for i in reader:
			roll=i[6]
			yag.send(i[1],attachments=path+"/"+roll+".xlsx")
			yag.send(i[4],attachments=path+"/"+roll+".xlsx")
			print("email sent to "+roll)

    


	


#generate_marksheet(r"C:\Users\Pranavi\python programming\sample_input\master_roll.csv",r"C:\Users\Pranavi\python programming\sample_input\responses.csv",5,-1)
#concise_marksheet(r"C:\Users\Pranavi\python programming\sample_input\master_roll.csv",r"C:\Users\Pranavi\python programming\sample_input\responses.csv",5,-1)
#Send_email(r"C:\Users\Pranavi\python programming\sample_input\master_roll.csv",r"C:\Users\Pranavi\python programming\sample_input\responses.csv",5,-1)
