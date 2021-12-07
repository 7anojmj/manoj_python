import csv
import os
os.system("cls")
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
directory=os.getcwd()
grade = {'AA':10,
         'AB':9,
         'BB':8,
         ' BB':8,
         'BC':7,
         'CC':6,
         'CD':5,
         'DD':4,
         'DD*':4,
         'F*':0,
         'F' :0,
         'I' :0
          }
def generate_marksheet():
    c=directory+"/output"
    try:
        os.makedirs(c)
    except:
        pass
    with open('names-roll.csv','r') as f1:
        reader1=csv.reader(f1)
        next(reader1)
        for x in reader1:
            wb = Workbook()
            wb.create_sheet(index=0, title="Overall")
            wb.create_sheet(index=1, title="Sem1")
            wb.create_sheet(index=2, title="Sem2")
            wb.create_sheet(index=3, title="Sem3")
            wb.create_sheet(index=4, title="Sem4")
            wb.create_sheet(index=5, title="Sem5")
            wb.create_sheet(index=6, title="Sem6")
            wb.create_sheet(index=7, title="Sem7")
            wb.create_sheet(index=8, title="Sem8")
            rno=str(x[0])
            head=['SI No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade']
            wb["Sem1"].append(head) 
            wb["Sem2"].append(head) 
            wb["Sem3"].append(head) 
            wb["Sem4"].append(head) 
            wb["Sem5"].append(head) 
            wb["Sem6"].append(head) 
            wb["Sem7"].append(head) 
            wb["Sem8"].append(head) 
            with open('grades.csv','r') as f2:
                reader2=csv.reader(f2)
                next(reader2)
                s1=s2=s3=s4=s5=s6=s7=s8=1
                c1=c2=c3=c4=c5=c6=c7=c8=0
                sp1=sp2=sp3=sp4=sp5=sp6=sp7=sp8=0
                tc1=tc2=tc3=tc4=tc5=tc6=tc7=tc8=0
                cp1=cp2=cp3=cp4=cp5=cp6=cp7=cp8=0
                for i in reader2:
                    if i[1]=='1' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s1,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c1=c1+int(i[3])
                                    s1=s1+1
                                    sp1=sp1+(int(i[3])*int(grade[i[4]]))
                                    wb["Sem1"].append(data)
                                tc1=c1
                                
                    elif i[1]=='2' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s2,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c2=c2+int(i[3])
                                    sp2=sp2+(int(i[3])*int(grade[i[4]]))
                                    s2=s2+1
                                    wb["Sem2"].append(data)
                                tc2=tc1+c2
                                
                    elif i[1]=='3' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s3,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c3=c3+int(i[3])
                                    sp3=sp3+(int(i[3])*int(grade[i[4]]))
                                    s3=s3+1
                                    wb["Sem3"].append(data)
                                tc3=tc2+c3
                                
                    elif i[1]=='4' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s4,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c4=c4+int(i[3])
                                    sp4=sp4+(int(i[3])*int(grade[i[4]]))
                                    s4=s4+1
                                    wb["Sem4"].append(data)
                                tc4=tc3+c4
                                
                    elif i[1]=='5' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s5,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c5=c5+int(i[3])
                                    sp5=sp5+(int(i[3])*int(grade[i[4]]))
                                    s5=s5+1
                                    wb["Sem5"].append(data)
                                tc5=tc4+c5
                               
                    elif i[1]=='6' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s6,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c6=c6+int(i[3])
                                    sp6=sp6+(int(i[3])*int(grade[i[4]]))
                                    s6=s6+1
                                    wb["Sem6"].append(data)
                                tc6=tc5+c6
                                
                    elif i[1]=='7' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s7,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c7=c7+int(i[3])
                                    sp7=sp7+(int(i[3])*int(grade[i[4]]))
                                    s7=s7+1
                                    wb["Sem7"].append(data)      
                                tc7=tc6+c7
                               
                    elif i[1]=='8' and i[0]==rno:
                        with open('subjects_master.csv','r') as f3:
                            reader3=csv.reader(f3)
                            next(reader3)
                            for j in reader3:
                                if i[2]==j[0]:
                                    data=[s8,i[2],j[1],j[2],i[3],i[5],i[4]]
                                    c8=c8+int(i[3])
                                    sp8=sp8+(int(i[3])*int(grade[i[4]]))
                                    s8=s8+1
                                    wb["Sem8"].append(data) 
                                tc8=tc7+c8 
                               
            sp1=round(sp1/c1,2)
            sp2=round(sp2/c2,2)
            if(c3):
                sp3=round(sp3/c3,2)
            if(c4):
                sp4=round(sp4/c4,2)
            if(c5):
                sp5=round(sp5/c5,2)
            if(c6):
                sp6=round(sp6/c6,2)
            if(c7):
                sp7=round(sp7/c7,2)
            if(c8):
                sp8=round(sp8/c8,2)  
            cp1=sp1
            cp2=((cp1*tc1)+(sp2*c2))/tc2
            if(c3):
                cp3=((cp2*tc2)+(sp3*c3))/tc3
            if(c4):
                cp4=((cp3*tc3)+(sp4*c4))/tc4
            if(c5):
                cp5=((cp4*tc4)+(sp5*c5))/tc5
            if(c6):
                cp6=((cp5*tc5)+(sp6*c6))/tc6
            if(c7):
                cp7=((cp6*tc6)+(sp7*c7))/tc7  
            if(c8):
                cp8=((cp7*tc7)+(sp8*c8))/tc8   
            wb["Overall"]['A1']='Roll No.'
            wb["Overall"]['B1']=rno
            wb["Overall"]['A2']='Name of Student ' 
            wb["Overall"]['B2']=x[1]
            wb["Overall"]['A3']='Discipline'
            wb["Overall"]['B3']=rno[4:6]
            l4=['Semester No.','1','2','3','4','5','6','7','8']
            wb["Overall"].append(l4)
            l5=['Semester wise Credit Taken',c1,c2,c3,c4,c5,c6,c7,c8]
            wb["Overall"].append(l5)
            l6=['SPI',sp1,sp2,sp3,sp4,sp5,sp6,sp7,sp8]
            wb["Overall"].append(l6)
            l7=['Total Credits Taken',tc1,tc2,tc3,tc4,tc5,tc6,tc7,tc8]
            wb["Overall"].append(l7)
            l8=['CPI',round(cp1,2),round(cp2,2),round(cp3,2),round(cp4,2),round(cp5,2),round(cp6,2),round(cp7,2),round(cp8,2)]
            wb["Overall"].append(l8)               
            wb.save(directory+'/'+'/output/'+rno+'.xlsx')
generate_marksheet()