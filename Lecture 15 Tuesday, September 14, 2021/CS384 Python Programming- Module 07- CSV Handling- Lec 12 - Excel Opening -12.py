# Openpyxl formulas
# We can write formula into the cell. These formulas are used to perform operations in excel file. After writing in the cell execute it from the workbook. Consider the following example:

from openpyxl import Workbook
wb = Workbook()
sheet = wb.active

rows_count = (
    (14, 27),
    (22, 30),
    (42, 92),
    (51, 32),
    (16, 60),
    (63, 13)
)

for i in rows_count:
    sheet.append(i)

cell = sheet.cell(row=7, column=3)
cell.value = "=AVERAGE(A1:B6)"
cell.font = cell.font.copy(bold=True)

cell = sheet.cell(row=8, column=3)
cell.value = "=sum(A1:B6)"
cell.font = cell.font.copy(bold=True)


cell = sheet.cell(row=7, column=3)
cell.value = "=AVERAGE(A1:B6)"
cell.font = cell.font.copy(bold=True)

cell = sheet.cell(row=9, column=3)
cell.value = "=max(A1:B6)"
cell.font = cell.font.copy(bold=True)


wb.save('formulas_book.xlsx')
