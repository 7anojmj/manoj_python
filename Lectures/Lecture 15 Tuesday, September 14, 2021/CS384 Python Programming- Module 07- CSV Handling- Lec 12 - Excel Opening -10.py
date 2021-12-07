from openpyxl import Workbook
# Openpyxl Iterate by rows
# The openpyxl provides the iter_row() function, which is used to read data corresponding to rows. Consider the following example:


wb = Workbook()
sheet = wb.active

rows = (
    (90, 46, 48, 44),
    (81, 30, 32, 16),
    (23, 95, 87, 27),
    (65, 12, 89, 53),
    (42, 81, 40, 44),
    (34, 51, 76, 42)
)

for row in rows:
    sheet.append(row)

for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()

wb.save('iter_rows.xlsx')
