# Reading All Values In The File


from openpyxl import load_workbook
wb = load_workbook("info.xlsx")
sheet = wb["Sheet1"]
row_count = sheet.max_row
column_count = sheet.max_column
print(row_count, column_count)

# exit(0)
for i in range(1, row_count + 1):
    for j in range(1, column_count + 1):
        data = sheet.cell(row=i, column=j).value
        print(data, end='   ')
    print('\n')
