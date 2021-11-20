from openpyxl import load_workbook
wb = load_workbook(r'sample_file.xlsx')

sheet = wb.active
sheet['A1'] = 'CS384'

sheet.cell(row=2, column=2).value = 5
wb.save(r'sample_file.xlsx')
